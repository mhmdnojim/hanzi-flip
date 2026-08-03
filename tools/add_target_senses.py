#!/usr/bin/env python3
"""Add a "Target Sense Map" sheet to a Sense Map / Reverse Index workbook.

Groups the Reverse Index by (Language, Main Entry) and gives each distinct
Chinese sense the entry maps to its own Target Sense ID, so the flashcard app
can build one clean card per target-language meaning in either direction.

Usage: python tools/add_target_senses.py HSK1.xlsx [out.xlsx]
"""
import re
import sys
from collections import OrderedDict, defaultdict

from openpyxl import load_workbook

SHEET = "Target Sense Map"
HEADERS = ["Language", "Main Entry", "Latin", "Target Sense ID",
           "Target Sense Gloss", "Linked Sense IDs", "Target Sense Note", "Coverage"]


def find_sheet(book, name):
    for s in book.sheetnames:
        if s.strip().lower() == name:
            return s
    return None


def rows(ws):
    it = ws.iter_rows(values_only=True)
    head = [str(h).strip() if h is not None else "" for h in next(it)]
    for r in it:
        yield {h: ("" if v is None else str(v).strip()) for h, v in zip(head, r)}


def code_for(language):
    letters = re.sub(r"[^A-Za-z]", "", language).upper()
    return (letters[:2] or "XX")


def main():
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else src
    book = load_workbook(src)

    rev_name = find_sheet(book, "reverse index")
    if not rev_name:
        sys.exit("No 'Reverse Index' sheet found.")

    # (language, entry) -> ordered map of gloss -> [sense ids], plus latin
    groups = OrderedDict()
    latins = {}
    for r in rows(book[rev_name]):
        lang, entry = r.get("Language", ""), r.get("Main Entry", "")
        sense_id = r.get("Sense ID", "")
        if not lang or not entry or not sense_id:
            continue
        key = (lang, entry)
        gloss = r.get("English Meaning", "").replace("|", ", ").strip()
        bucket = groups.setdefault(key, OrderedDict())
        bucket.setdefault(gloss, [])
        if sense_id not in bucket[gloss]:
            bucket[gloss].append(sense_id)
        latins.setdefault(key, r.get("Latin", ""))

    if SHEET in book.sheetnames:
        del book[SHEET]
    ws = book.create_sheet(SHEET)
    ws.append(HEADERS)

    counts = defaultdict(int)
    written = 0
    for (lang, entry), senses in groups.items():
        counts[(lang, entry)] = 0
        for gloss, ids in senses.items():
            counts[(lang, entry)] += 1
            n = counts[(lang, entry)]
            tid = f"{code_for(lang)}-{entry}-S{n:02d}"
            note = gloss if len(senses) > 1 else ""
            ws.append([lang, entry, latins.get((lang, entry), ""), tid, gloss,
                       "|".join(ids), note, "linked"])
            written += 1

    book.save(out)
    print(f"{written} target senses across {len(groups)} target words -> {out}")
    print("Append target-only meanings manually with an empty 'Linked Sense IDs' "
          "and Coverage = unlinked.")


if __name__ == "__main__":
    main()
